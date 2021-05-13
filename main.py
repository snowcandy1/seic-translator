# -- coding: utf-8 --
from trie import *

from openpyxl import load_workbook

import jieba

if __name__ == "__main__":
    wb = load_workbook("./translate.xlsx")
    sheet = wb["name"]
    trie = Trie()
    rows = sheet.max_row
    for i in range(2, rows + 1):
        trie.insert(sheet[i][0].value, sheet[i][1].value)
    # print(trie.root["共"])
    finder = TrieFinder(trie)
    writeout = load_workbook("./inputs.xlsx")
    wrsheet = writeout["names"]
    rows = wrsheet.max_row

    critical_list = []
    critical_wrd = []
    for i in range(2, rows + 1):
        gv = finder.split(wrsheet[i][0].value)
        # print(gv)
        wrsheet[i][1].value = gv
        seg = jieba.cut(wrsheet[i][0].value, cut_all=False)
        for word in seg:
            if len(word) > 1 and not trie.find(word) and word not in critical_wrd:
                critical_list.append([i, wrsheet[i][0].value, word])
                critical_wrd.append(word)

    wrsheet2 = writeout["critical"]
    i = 2
    last_critical = ""
    for critical in critical_list:
        wrsheet2[i][0].value = str(critical[0])
        if critical[1] != last_critical:
            wrsheet2[i][1].value = str(critical[1])
            last_critical = critical[1]
        wrsheet2[i][2].value = str(critical[2])
        i += 1

    writeout.save("./result.xlsx")


